import csv
import os

import openpyxl
from functools import wraps
from inspect import getfullargspec
from pathlib import Path

def to_path(fpath):
    if type(fpath) == str:
        fpath = Path(fpath)
    return fpath

def deco_fname_check(ftype):
    def _deco_fname_check(f):
        argspec = getfullargspec(f)
        argument_name = "fpath"
        argument_index = argspec.args.index(argument_name)
        @wraps(f)
        def wrapper(*args, **kwargs):
            if argument_index < len(args):
                fpath = to_path(args[argument_index])
            else:
                fpath = to_path(kwargs[argument_name])
            if fpath.suffix[1:] != ftype:
                msg = f"file name maybe wrong: expected ends with {ftype} but get {fpath}"
                print('\033[31m' + msg + '\033[0m') # ]] fix vim indent
            # do something with value
            return f(*args, **kwargs)
        return wrapper
    return _deco_fname_check


def load_as_list(fpath):
    fpath = to_path(fpath)
    if fpath.suffix == ".xlsx" and fpath.exists():
        return load_xlsx_as_list(fpath)
    else:
        fpath = os.path.splitext(fpath)[0] + ".csv"
        return load_csv_as_list(fpath)

# use `data_only=True` to get calculated value
def load_xlsx_as_list(fname, sheet_name=None, **kwargs):
    wb = openpyxl.load_workbook(fname, **kwargs)
    #print('wb.sheetnames', wb.sheetnames) # debug
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    return load_ws_as_list(ws)

# use `data_only=True` to get calculated value
def load_ws_as_list(ws):
    data = []
    for row in ws.iter_rows():
        one = []
        for cell in row:
            one.append(cell.value)
        data.append(one)
    return data

def load_csv_as_list(fpath, encoding="utf-16"):
    return list(csv.reader(open(fpath, encoding=encoding)))

def tie_key_value(header, value_list):
    data = []
    for row in value_list:
        new = {}
        for k, v in zip(header, row):
            new[k] = v
        data.append(new)
    return data


def load_as_dict(fpath):
    fpath = to_path(fpath)
    if fpath.suffix == ".xlsx" and fpath.exists():
        return load_xlsx_as_dict(fpath)
    else:
        fpath = os.path.splitext(fpath)[0] + ".csv"
        return load_csv_as_dict(fpath)

# use `data_only=True` to get calculated value
def load_xlsx_as_dict(fpath, sheet_name=None, **kwargs): #TODO rename
    wb = openpyxl.load_workbook(fpath, **kwargs)
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    return load_ws_as_dict(ws)

def load_ws_as_dict(ws):
    header = []
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            header.append(cell.value)
        break
    data = []
    for row in ws.iter_rows(min_row=2):
        one = {}
        for i, cell in enumerate(row):
            one[header[i]] = cell.value
        data.append(one)
    return data

def load_csv_as_dict(fpath, encoding=None):
    if encoding is None:
        encoding = guess_utf_encoding(fpath)
    reader = csv.DictReader(open(fpath, encoding=encoding))
    data = [row for row in reader]
    return data

def guess_utf_encoding(fpath):
    with open(fpath, "rb") as f:
        beginning = f.read(4)
        #NOTE The order of these if-statements is important. otherwise UTF32 LE may be detected as UTF16 LE as well.
        #NOTE delete -le -be to skip BOM.  ref. https://docs.python.org/ja/3.7/howto/unicode.html#reading-and-writing-unicode-data
        if beginning == b'\x00\x00\xfe\xff':
            return "utf-32"
            return "utf-32-be"
        elif beginning == b'\xff\xfe\x00\x00':
            return "utf-32"
            return "utf-32-le"
        elif beginning[0:3] == b'\xef\xbb\xbf':
            return "utf-8-sig"
        elif beginning[0:2] == b'\xff\xfe':
            return "utf-16"
            return "utf-16-le"
        elif beginning[0:2] == b'\xfe\xff':
            return "utf-16"
            return "utf-16-be"
        else:
            #NOTE handle unknown as utf-8
            return "utf-8"

def write_ll(fpath, data):
    fpath = to_path(fpath)
    if fpath.suffix == ".csv":
        write_csv_ll(fpath, data)
    else:
        write_xlsx_ll(fpath, data)

def write_dict(fpath, header, data):
    fpath = to_path(fpath)
    if fpath.suffix == ".csv":
        write_csv_dict(fpath, header, data)
    else:
        write_xlsx_dict(fpath, header, data)

@deco_fname_check("xlsx")
def write_xlsx_ll(fpath, data, style_func=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    write_ws_ll(ws, data, style_func)
    wb.save(fpath)
    wb.close()

def write_ws_ll(ws, data, style_func=None):
    for i, row in enumerate(data, 1):
        for j, value in enumerate(row, 1):
            write_cell(ws, i, j, value, style_func)

@deco_fname_check("csv")
def write_csv_ll(fpath, data):
    writer = csv.writer(open(fpath, "w", encoding="utf-16"))
    writer.writerows(data)

# data: list of dict
@deco_fname_check("xlsx")
def write_xlsx_dict(fpath, header, data, style_func=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws2 = wb.create_sheet(title="Sheet2")
    write_ws_dict(ws, header, data, style_func)
    wb.save(fpath)
    wb.close()

# data: list of dict
def write_ws_dict(ws, header, data, style_func):
    for j, key in enumerate(header, 1):
        ws.cell(row=1, column=j, value=key)
    for i, row in enumerate(data, 2):
        for j, key in enumerate(header, 1):
            value = row.get(key, None)
            write_cell(ws, i, j, value, style_func)

# use `kwargs={extrasaction: "ignore"}`
@deco_fname_check("csv")
def write_csv_dict(fpath, header, data, **kwargs):
    fpath = to_path(fpath)
    writer = csv.DictWriter(open(fpath, "w", encoding="utf-16"), header, **kwargs)
    writer.writeheader()
    for row in data:
        writer.writerow(row)

def write_cell(ws, i, j, value, style_func):
    cell = ws.cell(row=i, column=j, value=value)
    call_link = "https://call.ctrlq.org/"
    if type(value) is str and value.startswith(call_link):
        cell.hyperlink = value
        cell.value = value[len(call_link):]
    elif type(value) is str and (value.startswith("https://") or value.startswith("http://")) and " " not in value:
        cell.hyperlink = value
    if style_func is not None:
        style_func(cell)

from openpyxl.styles import PatternFill
ptn_fill_gray = PatternFill(patternType="solid", fgColor="999999")
def sample_style_func(cell):
    # ref. https://openpyxl.readthedocs.io/en/stable/api/openpyxl.cell.cell.html
    if cell.row % 2 == 0:
        cell.fill = ptn_fill_gray

def collect_header(data):
    s = set() # use set to check already registered
    header = [] # use list to save order
    for row in data:
        l = list(row.keys())
        for column_name in l:
            if column_name in s:
                continue
            s.add(column_name)
            header.append(column_name)
    return header

