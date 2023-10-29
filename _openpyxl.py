import openpyxl

def read(fpath):
    wb = openpyxl.load_workbook(fpath)
    print('wb.sheetnames', wb.sheetnames) # debug
    ws = wb["Sheet1"]
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            print('cell.value', cell.value) # debug
        break # for row is None

def read_as_dict_datum(fpath, sheet_name=None):
    wb = openpyxl.load_workbook(fpath)
    print('wb.sheetnames', wb.sheetnames) # debug
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    header = []
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            header.append(cell.value)
        break
    data = []
    for row in ws.iter_rows(min_row=2):
        one = {}
        for i, cell in enumerate(row):
            one[header[i]] = cell.value
        data.append(one)
        break # for row is None
    return data

# data: 2-dimension array
def write(fpath, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws2 = wb.create_sheet(title="Sheet2")
    for i, row in enumerate(data, 1):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)
    wb.save(fpath)
    wb.close()

# data: list of dict
def write_dict(fpath, header, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws2 = wb.create_sheet(title="Sheet2")
    for j, key in enumerate(header, 1):
        ws.cell(row=1, column=j, value=key)
    for i, row in enumerate(data, 2):
        for j, key in enumerate(header, 1):
            value = row.get(key, None)
            ws.cell(row=i, column=j, value=value)
    wb.save(fpath)
    wb.close()

def load_wb_from_csv(fpath):
    reader = csv.reader(open(fpath, encoding="utf_8_sig"))
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, row in enumerate(reader, 1):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)
    return wb

