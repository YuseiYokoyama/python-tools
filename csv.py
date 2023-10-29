import csv
import os

import openpyxl


def load_as_csv(fpath):
    if fpath.endswith(".xlsx") and os.path.exists(fpath):
        return read_xlsx(fpath)
    else:
        fpath = os.path.splitext(fpath)[0] + ".csv"
        return list(csv.reader(open(fpath, encoding="utf_8_sig")))

# use `data_only=True` to get calculated value 
def read_xlsx(fname, **kwargs):
    wb = openpyxl.load_workbook(fname, **kwargs)
    #print('wb.sheetnames', wb.sheetnames) # debug
    ws = wb.active
    return read_xlsx_ws(ws)

# use `data_only=True` to get calculated value 
def read_xlsx_ws(ws):
    data = []
    for row in ws.iter_rows():
        one = []
        for cell in row:
            one.append(cell.value)
        data.append(one)
    return data

def load_as_dict_csv(fpath):
    if fpath.endswith(".xlsx") and os.path.exists(fpath):
        return read_as_dict_datum(fpath)
    else:
        fpath = os.path.splitext(fpath)[0] + ".csv"
        return load_csv_as_dict(fpath)

# use `data_only=True` to get calculated value 
def read_as_dict_datum(fpath, sheet_name=None, **kwargs): #TODO rename
    wb = openpyxl.load_workbook(fpath, **kwargs)
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    return read_as_dict_datum_ws(ws) 

def read_as_dict_datum_ws(ws):
    header = []
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            header.append(cell.value)
        break
    data = []
    for row in ws.iter_rows(min_row=2):
        one = {}
        for i, cell in enumerate(row):
            one[header[i]] = cell.value
        data.append(one)
    return data

def load_csv_as_dict(fpath):
    reader = csv.DictReader(open(fpath, encoding="utf_8_sig"))
    data = [row for row in reader]
    return data

def write_csv_ll(fpath, data):
    writer = csv.writer(open(fpath, "w", encoding="utf_8_sig"))
    writer.writerows(data)

def write_ll_ws(ws, data):
    for i, row in enumerate(data, 1):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

# data: list of dict
def write_dict(fpath, header, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws2 = wb.create_sheet(title="Sheet2")
    write_dict_ws(ws, header, data)
    wb.save(fpath)
    wb.close()

# data: list of dict
def write_dict_ws(ws, header, data):
    for j, key in enumerate(header, 1):
        ws.cell(row=1, column=j, value=key)
    for i, row in enumerate(data, 2):
        for j, key in enumerate(header, 1):
            value = row.get(key, None)
            ws.cell(row=i, column=j, value=value)

def write_dict_to_csv(fpath, header, data):
    writer = csv.DictWriter(open(fpath, "w", encoding="utf_8_sig"), header)
    writer.writeheader()
    for row in data:
        writer.writerow(row)

