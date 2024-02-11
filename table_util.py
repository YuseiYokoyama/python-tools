import csv
import os

import openpyxl

def load_as_list(fpath):
    if fpath.endswith(".xlsx") and os.path.exists(fpath):
        return load_xlsx_as_list(fpath)
    else:
        fpath = os.path.splitext(fpath)[0] + ".csv"
        return list(csv.reader(open(fpath, encoding="utf-16")))

# use `data_only=True` to get calculated value 
def load_xlsx_as_list(fname, **kwargs):
    wb = openpyxl.load_workbook(fname, **kwargs)
    #print('wb.sheetnames', wb.sheetnames) # debug
    ws = wb.active
    return load_ws_as_list(ws)

# use `data_only=True` to get calculated value 
def load_ws_as_list(ws):
    data = []
    for row in ws.iter_rows():
        one = []
        for cell in row:
            one.append(cell.value)
        data.append(one)
    return data

def load_as_dict(fpath):
    if fpath.endswith(".xlsx") and os.path.exists(fpath):
        return load_xlsx_as_dict(fpath)
    else:
        fpath = os.path.splitext(fpath)[0] + ".csv"
        return load_csv_as_dict(fpath)

# use `data_only=True` to get calculated value 
def load_xlsx_as_dict(fpath, sheet_name=None, **kwargs): #TODO rename
    wb = openpyxl.load_workbook(fpath, **kwargs)
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    return load_ws_as_dict(ws) 

def load_ws_as_dict(ws):
    header = []
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            header.append(cell.value)
        break
    data = []
    for row in ws.iter_rows(min_row=2):
        one = {}
        for i, cell in enumerate(row):
            one[header[i]] = cell.value
        data.append(one)
    return data

def load_csv_as_dict(fpath):
    reader = csv.DictReader(open(fpath, encoding="utf-16"))
    data = [row for row in reader]
    return data

def write_xlsx_ll(fpath, data, style_func=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    write_ws_ll(ws, data, style_func)
    wb.save(fpath)
    wb.close()

def write_ws_ll(ws, data, style_func=None):
    for i, row in enumerate(data, 1):
        for j, value in enumerate(row, 1):
            write_cell(ws, i, j, value, style_func)

def write_csv_ll(fpath, data):
    writer = csv.writer(open(fpath, "w", encoding="utf-16"))
    writer.writerows(data)

# data: list of dict
def write_xlsx_dict(fpath, header, data, style_func=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws2 = wb.create_sheet(title="Sheet2")
    write_ws_dict(ws, header, data, style_func)
    wb.save(fpath)
    wb.close()

# data: list of dict
def write_ws_dict(ws, header, data, style_func):
    for j, key in enumerate(header, 1):
        ws.cell(row=1, column=j, value=key)
    for i, row in enumerate(data, 2):
        for j, key in enumerate(header, 1):
            value = row.get(key, None)
            write_cell(ws, i, j, value, style_func)

def write_csv_dict(fpath, header, data, **kwargs):
    writer = csv.DictWriter(open(fpath, "w", encoding="utf-16"), header, **kwargs)
    writer.writeheader()
    for row in data:
        writer.writerow(row)

def write_cell(ws, i, j, value, style_func):
    cell = ws.cell(row=i, column=j, value=value)
    if type(value) is str and (value.startswith("https://") or value.startswith("http://")) and " " not in value:
        cell.hyperlink = value
    if style_func is not None:
        style_func(cell)

